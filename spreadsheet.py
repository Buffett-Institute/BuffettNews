"""
Appends a new row to the Buffett News Page spreadsheet, matching the
existing sheet's columns and per-cell formatting exactly.
"""
import os

from copy import copy
from datetime import datetime

import openpyxl 
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

HEADER_LABELS = [
   "approved", "posted", "title", "date", "url", "source",
    "canva_title", "image_alt", "short_description", "content_type",
]
COLUMN_WIDTHS = [12, 10, 42, 14, 35, 25, 28, 45, 70, 20]
HEADER_ROW = 2


def build_spreadsheet(xlsx_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    #Define styles
    title_font = Font(name='Aptos Display', size=25, bold=True, color = "FF000000")
    title_fill = PatternFill(start_color="FF85B1F2", end_color="FF85B1F2", fill_type="solid")
    title_alignment = Alignment(horizontal="left")

    header_font = Font(name='Aptos Narrow', size=20, bold=True, color="FF000000")
    header_fill = PatternFill(start_color="FF8AACDE", end_color="FF8AACDE", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    #Add headers
    title = "Buffett News Page"
    ws.append([title])
    ws.append(HEADER_LABELS)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADER_LABELS))

    #Apply formatting to title
    for cell in ws[1]:
        cell.font = title_font
        cell.fill = title_fill
        cell.alignment = title_alignment
    #Apply formatting to headers
    for cell in ws[2]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    #Set column widths so text doesn't wrap into narrow columns
    for i, width in enumerate(COLUMN_WIDTHS, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    #Freeze title and header panes
    ws.freeze_panes = 'A3'

    os.makedirs(os.path.dirname(xlsx_path) or ".", exist_ok=True)
    wb.save(xlsx_path)

#Only build spreadsheet if it doesn't already exist
def _ensure_workbook(xlsx_path: str):
    if not os.path.exists(xlsx_path):
        build_spreadsheet(xlsx_path)

def _next_empty_row(ws) -> int:
    row = HEADER_ROW + 1
    while ws.cell(row=row, column=3).value not in (None, ""):
        row += 1
    return row

def append_row(xlsx_path: str, data: dict) -> int:
    """data keys: approved, posted, title, date (YYYY-MM-DD str or empty),
    url, source, canva_title, image_alt, short_description, content_type.
    Returns the row number written."""
    _ensure_workbook(xlsx_path)
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb.active
    row = _next_empty_row(ws)

    values = {
        1: data.get("approved", "") or "",
        2: data.get("posted", "") or "",
        3: data.get("title", "") or "",
        4: data.get("date") or "",
        5: data.get("url", "") or "",
        6: data.get("source", "") or "",
        7: data.get("canva_title", "") or "",
        8: data.get("image_alt", "") or "",
        9: data.get("short_description", "") or "",
        10: data.get("content_type", "") or "",
    }
    if values[4]:
        try:
            values[4] = datetime.strptime(values[4], "%Y-%m-%d")
        except ValueError:
            pass  # leave as free text if it doesn't parse

    body_font = Font(name='Aptos Narrow', size=20, bold=False, color="FF000000")
    body_alignment = Alignment(horizontal='left', wrap_text=True)
    thin_border = Side(border_style='thin', color='FF000000')
    border_style = Border( top=thin_border, left=thin_border, right=thin_border, bottom=thin_border)

    for col in range(1, 11):
        #src_cell = ws.cell(row=STYLE_SOURCE_ROW, column=col)
        dst_cell = ws.cell(row=row, column=col)
        dst_cell.value = values[col]
        dst_cell.font = body_font
        dst_cell.alignment = body_alignment
        dst_cell.border = border_style
        if col == 4: 
            dst_cell.number_format = "mm-dd-yy"
    wb.save(xlsx_path)
    return row


def read_recent_rows(xlsx_path: str, limit: int = 10):
    _ensure_workbook(xlsx_path)
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb.active
    last = HEADER_ROW

    for r in range(HEADER_ROW + 1, ws.max_row + 1):
        if ws.cell(row=r, column=3).value:
            last = r
    start = max(HEADER_ROW + 1, last - limit + 1)
    rows = []
    for r in range(start, last + 1):
        if not ws.cell(row=r, column=3).value:
            continue
        date_val = ws.cell(row=r, column=4).value
        rows.append({
            "row": r,
            "approved": ws.cell(row=r, column=1).value,
            "posted": ws.cell(row=r, column=2).value,
            "title": ws.cell(row=r, column=3).value,
            "date": date_val.strftime("%Y-%m-%d") if hasattr(date_val, "strftime") else (date_val or ""),
            "url": ws.cell(row=r, column=5).value,
            "source": ws.cell(row=r, column=6).value,
            "content_type": ws.cell(row=r, column=10).value,
        })
    rows.reverse()
    return rows
